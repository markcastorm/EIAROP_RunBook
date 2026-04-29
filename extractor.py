# extractor.py
# Parse the downloaded STEO_m.xlsx file: dynamically find the target tab,
# locate the papr_RS (Russia) row, map year/month headers, and extract
# all available periods into a structured list.

import logging

import openpyxl

import config

logger = logging.getLogger(__name__)


class STEOExtractor:
    """
    Extracts Russian oil production data from the EIA STEO workbook.

    Dynamically scans the workbook for the target tab, reads the header
    rows to build a year-month column map, finds the target row by code,
    and returns period-value pairs.
    """

    def __init__(self):
        self.logger = logger

    # ─────────────────────────────────────────────────────────────────────
    # Find the target tab
    # ─────────────────────────────────────────────────────────────────────

    def _find_target_tab(self, workbook):
        """
        Dynamically locate the target tab in the workbook.
        Returns the worksheet object.
        """
        target = config.TARGET_TAB
        sheet_names = workbook.sheetnames
        self.logger.info(f'Workbook tabs: {sheet_names}')

        # Exact match first
        if target in sheet_names:
            self.logger.info(f'Found target tab: {target}')
            return workbook[target]

        # Case-insensitive fallback
        for name in sheet_names:
            if name.lower() == target.lower():
                self.logger.info(f'Found target tab (case-insensitive): {name}')
                return workbook[name]

        raise RuntimeError(
            f'Target tab "{target}" not found in workbook. '
            f'Available tabs: {sheet_names}'
        )

    # ─────────────────────────────────────────────────────────────────────
    # Build column-to-period mapping from header rows
    # ─────────────────────────────────────────────────────────────────────

    def _build_column_map(self, worksheet):
        """
        Dynamically scan the worksheet to build a mapping: column_index → 'YYYY-MM'.

        Instead of assuming fixed row positions, we scan the first N rows to
        find:
          - The year-header row: the row with multiple integer values in the
            range 1900–2100 across different columns.
          - The month-header row: the row with month abbreviations (Jan–Dec)
            appearing in multiple columns.

        Returns:
            dict: {col_index: 'YYYY-MM', ...}
        """
        col_map = {}
        month_abbrs = set(config.MONTH_ABBR_TO_NUM.keys())

        # Scan the first 20 rows to find year and month header rows
        year_row_idx = None
        month_row_idx = None
        year_positions = {}  # {col_index: year_int}

        scan_limit = min(20, worksheet.max_row)

        for row_idx in range(1, scan_limit + 1):
            row = list(worksheet.iter_rows(
                min_row=row_idx, max_row=row_idx, values_only=False
            ))[0]

            # Count how many cells look like years (int 1900-2100)
            row_years = {}
            month_count = 0

            for cell in row:
                val = cell.value
                if val is None:
                    continue

                # Check for year
                try:
                    int_val = int(val)
                    if 1900 <= int_val <= 2100:
                        row_years[cell.column] = int_val
                except (ValueError, TypeError):
                    pass

                # Check for month abbreviation
                str_val = str(val).strip()
                if str_val in month_abbrs:
                    month_count += 1

            # A year-header row has multiple year values (at least 2)
            if len(row_years) >= 2 and year_row_idx is None:
                year_row_idx = row_idx
                year_positions = row_years
                self.logger.info(
                    f'Year header row found at row {row_idx}: {row_years}'
                )

            # A month-header row has many month abbreviations (at least 6)
            if month_count >= 6 and month_row_idx is None:
                month_row_idx = row_idx
                self.logger.info(
                    f'Month header row found at row {row_idx} '
                    f'({month_count} month labels)'
                )

            # Stop scanning once both are found
            if year_row_idx is not None and month_row_idx is not None:
                break

        if not year_positions:
            raise RuntimeError(
                f'No year header row found in the first {scan_limit} rows'
            )
        if month_row_idx is None:
            raise RuntimeError(
                f'No month header row found in the first {scan_limit} rows'
            )

        # Re-read the month row to build the full column map
        month_row = list(worksheet.iter_rows(
            min_row=month_row_idx, max_row=month_row_idx, values_only=False
        ))[0]

        sorted_year_cols = sorted(year_positions.keys())

        for cell in month_row:
            month_abbr = str(cell.value).strip() if cell.value else ''
            if month_abbr not in config.MONTH_ABBR_TO_NUM:
                continue

            col = cell.column
            month_num = config.MONTH_ABBR_TO_NUM[month_abbr]

            # The year for a column is the largest year_col that is <= this column
            assigned_year = None
            for year_col in sorted_year_cols:
                if year_col <= col:
                    assigned_year = year_positions[year_col]
                else:
                    break

            if assigned_year is not None:
                period = f'{assigned_year}-{month_num}'
                col_map[col] = period

        self.logger.info(f'Column map built: {len(col_map)} periods '
                         f'({min(col_map.values())} to {max(col_map.values())})')
        return col_map

    # ─────────────────────────────────────────────────────────────────────
    # Find the target data row
    # ─────────────────────────────────────────────────────────────────────

    def _find_target_row(self, worksheet):
        """
        Scan all rows to find the one where column A matches TARGET_ROW_CODE.
        Returns the row number (1-indexed).
        """
        target_code = config.TARGET_ROW_CODE

        for row_idx, row in enumerate(worksheet.iter_rows(values_only=False), 1):
            cell_a = row[0]  # Column A
            if cell_a.value is not None:
                cell_val = str(cell_a.value).strip()
                if cell_val == target_code:
                    # Verify sublabel in column B
                    cell_b = row[1] if len(row) > 1 else None
                    sublabel = str(cell_b.value).strip() if cell_b and cell_b.value else ''
                    self.logger.info(
                        f'Found target row {row_idx}: '
                        f'code="{cell_val}", sublabel="{sublabel}"'
                    )
                    return row_idx

        raise RuntimeError(
            f'Target row "{target_code}" not found in worksheet'
        )

    # ─────────────────────────────────────────────────────────────────────
    # Extract data for the target row
    # ─────────────────────────────────────────────────────────────────────

    def _extract_row_data(self, worksheet, row_num, col_map):
        """
        Read all data values from the target row, mapped to periods.

        Returns:
            list of (period, value) tuples sorted by period.
            e.g. [('2022-01', 11.2776), ('2022-02', 11.3308), ...]
        """
        data = []
        row = list(worksheet.iter_rows(
            min_row=row_num, max_row=row_num, values_only=False
        ))[0]

        for cell in row:
            col = cell.column
            if col not in col_map:
                continue

            period = col_map[col]
            value = cell.value

            if value is not None:
                try:
                    numeric_value = float(value)
                    data.append((period, numeric_value))
                except (ValueError, TypeError):
                    self.logger.warning(
                        f'Non-numeric value at col {col} ({period}): {value}'
                    )

        data.sort(key=lambda x: x[0])
        self.logger.info(f'Extracted {len(data)} data points')
        return data

    # ─────────────────────────────────────────────────────────────────────
    # Public API
    # ─────────────────────────────────────────────────────────────────────

    def extract(self, file_path):
        """
        Parse the STEO workbook and extract papr_RS (Russia) data.

        Args:
            file_path: path to the downloaded STEO_m.xlsx

        Returns:
            list of (period, value) tuples:
            [('2022-01', 11.2776), ('2022-02', 11.3308), ...]
        """
        self.logger.info(f'Opening workbook: {file_path}')
        wb = openpyxl.load_workbook(file_path, data_only=True)

        # Step 1: Find the target tab
        ws = self._find_target_tab(wb)

        # Step 2: Build column → period mapping from headers
        col_map = self._build_column_map(ws)

        # Step 3: Find the target row (papr_RS)
        row_num = self._find_target_row(ws)

        # Step 4: Extract data values
        data = self._extract_row_data(ws, row_num, col_map)

        wb.close()
        return data
